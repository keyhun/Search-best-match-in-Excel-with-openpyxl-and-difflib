import openpyxl
from difflib import get_close_matches

# 엑셀 열기
workbook = openpyxl.load_workbook('mapping_test.xlsx')
sheet = workbook.active

# A열 데이터 읽기
a_column = [cell.value for cell in sheet['A'] if cell.value is not None]

# B열 데이터 읽기 및 C열, D열에 결과 저장
for row in range(1, sheet.max_row + 1):  # 첫행에 헤더가 있을 경우 1->2로 변경 
    b_value = sheet.cell(row=row, column=2).value  
    if b_value:
        matches = get_close_matches(b_value, a_column, n=1, cutoff=0.6)  
        closest_match = matches[0] if matches else None
        
        # C열에 가장 유사한 문자열 저장
        sheet.cell(row=row, column=3, value=closest_match)
        
        # D열에 가장 유사한 문자열의 엑셀 좌표 저장
        if closest_match:
            a_row_index = a_column.index(closest_match) + 1  
            cell_address = f'A{a_row_index}'  
            sheet.cell(row=row, column=4, value=cell_address)

workbook.save('mapping_test_outcome.xlsx')
